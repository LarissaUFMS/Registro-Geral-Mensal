from turtle import color
import pyodbc
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles.numbers import NumberFormatList
from openpyxl.styles import Font, Fill
from openpyxl.styles import PatternFill


def nomeWorkBook(nomeR):
    if ('BOLSAO/PARANAIBA' in nomeR):
        return 'registroPNB.xlsx'
    elif ('GRANDE DOURADOS' in nomeR):
        return 'registroDOS.xlsx'
    elif ('NORTE' in nomeR):
        return 'registroCXM.xlsx'
    elif ('JIM' in nomeR):
        return 'registroJIM.xlsx'
    elif ('CONE-SUL' in nomeR):
        return 'registroNVR.xlsx'
    elif ('SUL/FRONTEIRA' in nomeR):
        return 'registroPPR.xlsx'
    elif ('LESTE' in nomeR):
        return 'registroNDI.xlsx'
    elif ('PANTANAL/CORUMBA' in nomeR):
        return 'registroCMA.xlsx'
    elif ('PANTANAL/AQUIDAUANA' in nomeR):
        return 'registroAUA.xlsx'
    elif ('BOLSAO/TRES LAGOAS' in nomeR):
        return 'registroTLS.xlsx'

def plotGraph(ws, max_row, nomeL):
	
	#Gráfico RDA e RCE
	chart1 = ScatterChart()
	#Gráfico Econ e DIFF
	chart2 = ScatterChart()

	xvalues = Reference(ws, min_col = 1, min_row = 2, max_row = max_row)

	yvalues = Reference(ws, min_col = 2, min_row = 2, max_row = max_row)
	series = Series(yvalues, xvalues, title = 'RDA')
	chart1.series.append(series)

	yvalues = Reference(ws, min_col = 3, min_row = 2, max_row = max_row)
	series = Series(yvalues, xvalues, title = 'RCE')
	chart1.series.append(series)

	yvalues = Reference(ws, min_col = 4, min_row = 2, max_row = max_row)
	series = Series(yvalues, xvalues, title = 'ECON')
	chart2.series.append(series)

	yvalues = Reference(ws, min_col = 5, min_row = 2, max_row = max_row)
	series = Series(yvalues, xvalues, title = 'DIFF')
	chart2.series.append(series)
	
	chart1.title = nomeL + '--RDA e RCE'
	ws.add_chart(chart1, 'G3')
	chart2.title = nomeL + '--ECON e DIFF'
	ws.add_chart(chart2, 'G18')
    
if __name__ == '__main__':
    # Conectando com o banco de dados
    con = pyodbc.connect(
        'DRIVER={SQL Server};SERVER=10.100.100.48\\SCI;PORT=1433;DATABASE=SCI;Trusted_Connection=yes;')
    # Consultando a tabela do banco
    conRegional = 'select distinct(regional) from resultado;'
    cursor = con.cursor()
    cursor.execute(conRegional)
    regional = cursor.fetchall()
    for r in regional:
        conLocal = 'select distinct(local) from resultado where regional =? order by local'
        cursor.execute(conLocal, r[0])
        local = cursor.fetchall()
        wb = Workbook()
        for l in local:
            ws = wb.create_sheet(l[0], -1)
            conResultado = """select datageracao, referencia AS DATA, RDA, RCE, ECON, DDIFF from historico_resultado
			WHERE local = ?
			union
			select datageracao, referencia AS DATA, RDA, RCE, ECON, DDIFF from resultado where datageracao in (select max(DATAGERACAO) as DATAGERACAO from resultado group by referencia, local)
			and local = ?
			order by datageracao;"""
            col = ['DATA', 'RDA', 'RCE', 'ECON', 'DDIFF']
            ws.append(col)
            
            nLinhas = 0
            cursor.execute(conResultado, l[0], l[0])
            row = cursor.fetchone()
            anterior = 0
            wsAnteriorB = True
            wsAnteriorC = True
            wsAnteriorD = True
            b = 0
            c = 0
            d = 0
            while (row != None):
                row[1] = datetime.strptime(row[1], '%m/%Y')
                if(row[2] != anterior):
                    nLinhas += 1
                    ws.append(row[1:])
                    if(ws['B'+str(nLinhas + 1)].value and wsAnteriorB):
                        b = nLinhas + 1
                        wsAnteriorB = False
                    if(ws['C'+str(nLinhas + 1)].value and wsAnteriorC):
                        c = nLinhas + 1
                        wsAnteriorC = False
                    if(ws['D'+str(nLinhas + 1)].value and wsAnteriorD):
                        d = nLinhas + 1
                        wsAnteriorD = False
                row = cursor.fetchone()
            for cell in ws['A']:
                cell.number_format = 'mmm-yy'
            plotGraph(ws, nLinhas + 1, l[0])
            saldo =str(nLinhas + 2)
            
            ws['A' + saldo] = 'Saldo = '
            ws['A' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
            ws['B' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
            ws['C' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
            ws['D' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
            ws['E' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
            
            if(b > 0):
                ws['B' + saldo].value = ws['B' + str(nLinhas + 1)].value - ws['B' + str(b)].value
            else:
                ws['B' + saldo].value = 0
            if(c > 0):
                if(ws['C' + str(nLinhas + 1)].value == None):
                    ws['C' + str(nLinhas + 1)].value = 0
                ws['C' + saldo].value = ws['C' + str(nLinhas + 1)].value - ws['C' + str(c)].value  
            else:
                ws['C' + saldo].value = 0
            if(d > 0):    
                ws['D' + saldo].value = ws['D' + str(nLinhas + 1)].value - ws['D' + str(d)].value
                ws['E' + saldo].value = ws['E' + str(nLinhas + 1)].value - ws['E' + str(d)].value
            else:
                ws['D' + saldo].value = 0
                ws['E' + saldo].value = 0
                
        nomeWB = nomeWorkBook(r[0])
        wb.save(nomeWB)
