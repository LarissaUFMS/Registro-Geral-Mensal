#pyinstaller RegistroRegional.py --onefile --noconsole

from turtle import color
import pyodbc
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles.numbers import NumberFormatList
from openpyxl.styles import Font, Fill
from openpyxl.styles import PatternFill



def plotGraph(ws, max_row, nomeL):

	# Gráfico RDA e RCE
	chart1 = ScatterChart()
	# Gráfico Econ e DIFF
	chart2 = ScatterChart()

	xvalues = Reference(ws, min_col=1, min_row=2, max_row=max_row)

	yvalues = Reference(ws, min_col=2, min_row=2, max_row=max_row)
	series = Series(yvalues, xvalues, title='RDA')
	chart1.series.append(series)

	yvalues = Reference(ws, min_col=3, min_row=2, max_row=max_row)
	series = Series(yvalues, xvalues, title='RCE')
	chart1.series.append(series)

	yvalues = Reference(ws, min_col=4, min_row=2, max_row=max_row)
	series = Series(yvalues, xvalues, title='ECON')
	chart2.series.append(series)

	yvalues = Reference(ws, min_col=5, min_row=2, max_row=max_row)
	series = Series(yvalues, xvalues, title='DIFF')
	chart2.series.append(series)

	chart1.title = nomeL + '--RDA e RCE'
	ws.add_chart(chart1, 'G3')
	chart2.title = nomeL + '--ECON e DIFF'
	ws.add_chart(chart2, 'G18')
 
def nome(nomeR):
    if ('BOLSAO/PARANAIBA' in nomeR):
        return 'PNB'
    elif ('GRANDE DOURADOS' in nomeR):
        return 'DOS'
    elif ('NORTE' in nomeR):
        return 'CXM'
    elif ('JIM' in nomeR):
        return 'JIM'
    elif ('CONE-SUL' in nomeR):
        return 'NVR'
    elif ('SUL/FRONTEIRA' in nomeR):
        return 'PPR'
    elif ('LESTE' in nomeR):
        return 'NDI'
    elif ('PANTANAL/CORUMBA' in nomeR):
        return 'CMA'
    elif ('PANTANAL/AQUIDAUANA' in nomeR):
        return 'AUA'
    elif ('BOLSAO/TRES LAGOAS' in nomeR):
        return 'TLS'


if __name__ == '__main__':
    # Conectando com o banco de dados
    con = pyodbc.connect(
        'DRIVER={SQL Server};SERVER=10.100.100.48\\SCI;PORT=1433;DATABASE=SCI;Trusted_Connection=yes;')
    # Consultando a tabela do banco
    conRegional = 'select distinct(regional) from resultado;'
    cursor = con.cursor()
    cursor.execute(conRegional)
    regional = cursor.fetchall()
    wb = Workbook()
    conResultado = """select regional, max(datageracao) as datageracao, referencia as DATA, sum(RDA) as RDA, sum(RCE) as RCE, sum(ECON) as ECON, sum(DDIFF) as DDIFF  from historico_resultado
			where regional = ? group by regional, referencia
			union
			select regional, max(datageracao) as datageracao, referencia as DATA, sum(RDA) as RDA, sum(RCE) as RCE, sum(ECON) as ECON, sum(DDIFF) as DDIFF from resultado
			where datageracao in (select max(DATAGERACAO) as DATAGERACAO from resultado group by referencia, local)
			and regional = ?
			group by regional, referencia order by regional, datageracao;"""
    for r in regional:  
        ws = wb.create_sheet(nome(r[0]), -1)
        col = ['DATA', 'RDA', 'RCE', 'ECON', 'DDIFF']
        ws.append(col)
        nLinhas = 0
        cursor.execute(conResultado, r[0], r[0])
        row = cursor.fetchone()
        anterior = 0
        wsAnterior = True
        while (row != None):
            row[2] = datetime.strptime(row[2], '%m/%Y')
            if(row[2] != anterior):
                nLinhas += 1
                ws.append(row[2:])
            anterior = row[2]
            row = cursor.fetchone()
            if(ws['D'+str(nLinhas + 1)].value and wsAnterior):
                f = str(nLinhas + 1)
                wsAnterior = False
                
        for cell in ws['A']:
            cell.number_format = 'mmm-yy'
        plotGraph(ws, nLinhas +1 , r[0])
        saldo =str(nLinhas + 2)
        ws['A' + saldo] = 'Saldo = '
        ws['A' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
       
        ws['B' + saldo].value = ws['B' + str(nLinhas + 1)].value - ws['B2'].value
        ws['B' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

        ws['C' + saldo].value = ws['C' + str(nLinhas + 1)].value - ws['C2'].value
        ws['C' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

        ws['D' + saldo].value = ws['D' + str(nLinhas + 1)].value - ws['D' + f].value
        ws['D' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

        ws['E' + saldo].value = ws['E' + str(nLinhas + 1)].value - ws['E' + f].value
        ws['E' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
    wb.save('registroRegional.xlsx')
