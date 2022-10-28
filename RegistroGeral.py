# Conector de sql recomendado na documentação oficial.
from turtle import color
import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles.numbers import NumberFormatList
from openpyxl.styles import Font, Fill
from openpyxl.styles import PatternFill
#pyinstaller --onefile RegistroGeral.py
#from joblib import Parallel, delayed

def createWorkBook(nomeR, regional):
	if('BOLSAO/PARANAIBA' in nomeR):
		createSheetLocal('PNB', regional)
	elif('GRANDE DOURADOS' in nomeR):
		createSheetLocal('DOS', regional)
	elif('NORTE' in nomeR):
		createSheetLocal('CXM', regional)
	elif('JIM' in nomeR):
		createSheetLocal('JIM', regional)
	elif('CONE-SUL' in nomeR):
		createSheetLocal('NVR', regional)
	elif('SUL/FRONTEIRA'in nomeR):
		createSheetLocal('PPR', regional)
	elif('LESTE' in nomeR):
		createSheetLocal('NDI', regional)
	elif('PANTANAL/CORUMBA' in nomeR):
		createSheetLocal('CMA', regional)
	elif('PANTANAL/AQUIDAUANA' in nomeR):
		createSheetLocal('AUA', regional)

def createSheetLocal(nomeR, regional):
	wb = Workbook()
	#Listas de colunas necessárias para abrir no excel
	col = ['DATA', 'RDA', 'RCE', 'ECON', 'DDIFF']
	groupLocal = regional.groupby('LOCALIDADE')
	for nomeL, localidade in groupLocal:
		#graphImg = plotGraph(nameL, localidade)
		ws = wb.create_sheet(nomeL, -1)
		#ws.add_image(graphImg, 'F2')
		for r in dataframe_to_rows(localidade[col], index = False, header = True):
			ws.append(r)
		NumberFormatList()
		for cell in ws['A']:
			cell.number_format = 'mmm-yy'
			#cell.number_format = builtin_format_code(17)
		n = len(localidade) + 1
		row_n = str(n)
		saldo = str(n+1)

		if(ws['B2'].value):
			ws['B2'].value = 0
		if(ws['C2'].value):
			ws['C2'].value = 0
		if(ws['D2'].value):
			ws['D2'].value = 0
		if(ws['E2'].value):
			ws['E2'].value = 0

		ws['A' + saldo] = 'Saldo = '
		ws['A' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
		ws['B' + saldo].value = ws['B' + row_n].value - ws['B2'].value
		ws['B' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
		ws['C' + saldo].value = ws['C' + row_n].value - ws['C2'].value
		ws['C' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
		ws['D' + saldo].value = ws['D' + row_n].value - ws['D2'].value
		ws['D' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
		ws['E' + saldo].value = ws['E' + row_n].value - ws['E2'].value
		ws['E' + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

		plotGraph(ws, localidade, nomeL)

	wb.save('registro' + nomeR +'.xlsx')

#def plotGraph(nameL, localidade):
	#plt.plot(localidade['REFERENCIA'], localidade['RDA'], linestyle='--', color='b', linewidth=3.0)
	#nColumns = len(localidade)
	#listOf_Xticks = localidade['REFERENCIA'].iloc[list(range(0, nColumns, 6))]
	#print(listOf_Xticks)
	#plt.xticks(listOf_Xticks) 
	#plt.title(nameL)
	#imgPath = nameL + 'grafico.jpg'
	#plt.savefig(imgPath)
	#graphImage = Image(imgPath)
	#return graphImage
def plotGraph(ws, localidade, nomeL):
	max_row = len(localidade) + 1
	#Gráfico RDA e RCE
	chart1 = ScatterChart()
	#Gráfico Econ e DIFF
	chart2 = ScatterChart()

	#chartRDA = ScatterChart()
	#chartRCE = ScatterChart()
	#chartECON = ScatterChart()
	#chartDIFF = ScatterChart()

	xvalues = Reference(ws, min_col = 1, min_row = 2, max_row = max_row)

	yvalues = Reference(ws, min_col = 2, min_row = 2, max_row = max_row)
	series = Series(yvalues, xvalues, title = 'RDA')
	chart1.series.append(series)
	#chartRDA.series.append(series)

	yvalues = Reference(ws, min_col = 3, min_row = 2, max_row = max_row)
	series = Series(yvalues, xvalues, title = 'RCE')
	chart1.series.append(series)
	#chartRCE.series.append(series)

	yvalues = Reference(ws, min_col = 4, min_row = 2, max_row = max_row)
	series = Series(yvalues, xvalues, title = 'ECON')
	chart2.series.append(series)
	#chartECON.series.append(series)

	yvalues = Reference(ws, min_col = 5, min_row = 2, max_row = max_row)
	series = Series(yvalues, xvalues, title = 'DIFF')
	chart2.series.append(series)
	#chartDIFF.series.append(series)
	
	chart1.title = nomeL + '--RDA e RCE'
	ws.add_chart(chart1, 'G3')
	chart2.title = nomeL + '--ECON e DIFF'
	ws.add_chart(chart2, 'G18')

	#chartRDA.title = nomeL +'-RDA'
	#ws.add_chart(chartRDA, 'F18')
	
	#chartRCE.title = nomeL +'-RCE'
	#ws.add_chart(chartRCE, 'F33')
	
	#chartECON.title = nomeL +'-ECON'
	#ws.add_chart(chartECON, 'O3')
	
	#chartDIFF.title = nomeL +'-DIFF'
	#ws.add_chart(chartDIFF, 'O18')

if __name__ == '__main__':
	# Conectando com o banco de dados
	con = pyodbc.connect(
	'DRIVER={SQL Server};SERVER=10.100.100.48\\SCI;PORT=1433;DATABASE=SCI;Trusted_Connection=yes;')
	#Consultando a tabela do banco
	conHistorico= 'select REGIONAL, LOCAL as LOCALIDADE, REFERENCIA, RDA, RCE, ECON, DDIFF, DATAGERACAO as DATA from dbo.historico_resultado ORDER BY REGIONAL, LOCAL, DATAGERACAO asc'
	conResultado ='select REGIONAL, LOCAL as LOCALIDADE, REFERENCIA, RDA, RCE, ECON, DDIFF, DATAGERACAO as DATA from dbo.resultado ORDER BY REGIONAL, LOCAL, DATAGERACAO asc;'
	#Convertendo em REFERENCIAframe
	dfHistorico = pd.read_sql(conHistorico, con)
	dfResultado = pd.read_sql(conResultado, con)
	#Agrupando por Localidade e REFERENCIA
	dfgroup = dfResultado.groupby(["LOCALIDADE", "REFERENCIA"])
	#Selecionando a primeira linha de cada grupo
	dfResultado = dfgroup.tail(1)
	#Concatenando os REFERENCIAframes
	df = pd.concat([dfHistorico, dfResultado], ignore_index=True, sort=False)
	#Apagando as linhas duplicadas
	df = df.drop_duplicates(subset=['REFERENCIA', 'LOCALIDADE'], keep='last')
	#Convertendo o formato da coluna REFERENCIA que estava em string para datetime
	#df['REFERENCIA'] = pd.to_datetime(df['REFERENCIA'], format='%m/%Y')

	#Ordenando o REFERENCIAframe primeiro pela redional e depois por localidade
	df = df.sort_values(by = ['REGIONAL', 'LOCALIDADE'])
	#Agrupando por Regional
	groupRegional = df.groupby("REGIONAL")
	# Create a new directory if not exist
	#if not os.path.exists('Graficos'):
	#	os.makedirs('Graficos')	
	#Iterar no grupo para criar vários worbooks por regional.
	
	for nomeR, regional in groupRegional:
		createWorkBook(nomeR, regional)